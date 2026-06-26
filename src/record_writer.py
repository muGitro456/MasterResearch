import numpy as np
import pandas as pd
import os
import openpyxl
from openpyxl.styles.alignment import Alignment

import logger


def write4plot(trial: int, nums: tuple, f_name: str, m_name: str, s_time) -> str:
    dir_path = (
        '../backLog/'
        + nums[0] + '_' + m_name + '/'
        + nums[1] + '_' + f_name + '/'
        + s_time.strftime('%Y%m%d_%H%M%S') + '/'
    )
    os.makedirs(dir_path, exist_ok=True)

    col = ['f1', 'f2', 'f3'] if f_name == "DTLZ1" else ['f1', 'f2']
    row = [str(r) for r in range(logger.LOG_FIT[-1].shape[0])]
    df = pd.DataFrame(logger.LOG_FIT[-1], index=row, columns=col)
    df.to_csv(os.path.join(dir_path, 'front_' + nums[0] + nums[1] + '_' + str(trial).zfill(3)) + '.csv')

    return dir_path


def write_record(sheet_name: str, trial: int, start_time, names: tuple, comment: str,
                 processing_time: float, n_sub: int, *indicators) -> None:
    my_wb = openpyxl.load_workbook(sheet_name)
    my_sheet = my_wb['No.7']
    right_alignment = Alignment(horizontal='right', vertical='center')

    r = 1
    while True:
        if my_sheet.cell(row=r, column=1).value is None:
            break
        else:
            r += 1

    my_sheet.cell(r, 1).value = start_time.year
    my_sheet.cell(r, 2).value = start_time.month
    my_sheet.cell(r, 3).value = start_time.day
    my_sheet.cell(r, 4).value = start_time.strftime('%H:%M:%S')
    my_sheet.cell(r, 5).value = names[0]
    my_sheet.cell(r, 6).value = names[1]
    my_sheet.cell(r, 7).value = names[2]
    my_sheet.cell(r, 8).value = trial
    my_sheet.cell(r, 9).value = comment
    my_sheet.cell(r, 10).value = processing_time
    my_sheet.cell(r, 11).value = n_sub

    for i, indicator in enumerate(indicators):
        my_sheet.cell(r, 12 + 6*i).value = np.average(indicator)
        my_sheet.cell(r, 13 + 6*i).value = np.max(indicator)
        my_sheet.cell(r, 14 + 6*i).value = np.min(indicator)
        my_sheet.cell(r, 15 + 6*i).value = np.median(indicator)
        my_sheet.cell(r, 16 + 6*i).value = "No." + str(np.argmax(indicator) + 1)
        my_sheet.cell(r, 17 + 6*i).value = "No." + str(np.argmin(indicator) + 1)

    my_sheet.cell(r, 4).alignment = right_alignment
    my_wb.save(sheet_name)
    print("Save Successed!")
