import numpy as np
import pandas as pd
import os, sys, glob, openpyxl
from openpyxl.styles.alignment import Alignment

LOG_POS = []
LOG_VEL = []
LOG_FIT = []

def store(var, target):
    match target:
        case 'p':
            LOG_POS.append(var)
        case 'v':
            LOG_VEL.append(var)
        case 'e':
            LOG_FIT.append(var)
        case _:
            print("該当なし")
            return -1

def write4Debug(target, maxgen, mNum, mName):
    # 新しいディレクトリの作成
    new_dir_path = '../backLog/' + mNum + '_' + mName + '/'
    os.makedirs(new_dir_path, exist_ok=True)

    match target:
        case 'p':  # 位置(position)の場合
            col = ['x' + str(d) for d in range(LOG_POS[0].shape[1])]
            row = [str(d % LOG_POS[0].shape[0]) for d in range(LOG_POS[0].shape[0] * maxgen)]
            tmp = np.array(LOG_POS) # listをnumpy.ndarrayに変換
            tmp = (tmp.flatten()).reshape(LOG_POS[0].shape[0] * maxgen, LOG_POS[0].shape[1]).copy() # copy()を付けて値を反映!

            df = pd.DataFrame(tmp, index=row, columns=col)
            df.to_csv(os.path.join(new_dir_path,  mName + '_pos_SP.csv'), header=True, index=True)
        
        case 'v':  # 速度(velocity)の場合
            col = ['v' + str(d) for d in range(LOG_VEL[0].shape[1])]
            row = [str(d % LOG_VEL[0].shape[0]) for d in range(LOG_VEL[0].shape[0] * maxgen)]
            tmp = np.array(LOG_VEL) # listをnumpy.ndarrayに変換
            tmp = (tmp.flatten()).reshape(LOG_VEL[0].shape[0] * maxgen, LOG_VEL[0].shape[1]).copy() # copy()を付けて値を反映!

            df = pd.DataFrame(tmp, index=row, columns=col)
            df.to_csv(os.path.join(new_dir_path, mName + '_vel_SP.csv'), header=True, index=True)
        
        case _:  # デフォルト処理
            print("Error : database")
            return -1

def write4Plot(trial, nums, fName, mName, sTime):
    global new_dir_path_log
    new_dir_path_log = '../backLog/' + nums[0] + '_' + mName + '/' + nums[1] + '_' + fName + '/' + sTime.strftime('%Y%m%d_%H%M%S') + '/'
    os.makedirs(new_dir_path_log, exist_ok=True)

    if fName == "DTLZ1":
        col = ['f1', 'f2', 'f3']
    else:
        col = ['f1', 'f2']

    row = [str(r) for r in range(LOG_FIT[-1].shape[0])]
    df = pd.DataFrame(LOG_FIT[-1], index=row, columns=col)
    df.to_csv(os.path.join(new_dir_path_log, 'front_' + nums[0] + nums[1] + '_' + str(trial).zfill(3)) + '.csv')

def write_record(sheet_name: str, TRIAL: int, startTime, names: tuple, comment: str, processingTime, N_SUB, *indicators):
    my_wb = openpyxl.load_workbook(sheet_name)
    my_sheet = my_wb['No.7']
    right_alignment = Alignment(horizontal='right',vertical='center') #右揃えの定義

    # 未記入の行を探査
    r = 1
    while True:
        if my_sheet.cell(row=r, column=1).value is None:
            break
        else:
            r += 1
    
    my_sheet.cell(r, 1).value = startTime.year
    my_sheet.cell(r, 2).value = startTime.month
    my_sheet.cell(r, 3).value = startTime.day
    my_sheet.cell(r, 4).value = startTime.strftime('%H:%M:%S')
    my_sheet.cell(r, 5).value = names[0]
    my_sheet.cell(r, 6).value = names[1]
    my_sheet.cell(r, 7).value = names[2]
    my_sheet.cell(r, 8).value = TRIAL
    my_sheet.cell(r, 9).value = comment
    my_sheet.cell(r, 10).value = processingTime
    my_sheet.cell(r, 11).value = N_SUB

    for i, indicator in enumerate(indicators):
        my_sheet.cell(r, 12 + 6*i).value = np.average(indicator)
        my_sheet.cell(r, 13 + 6*i).value = np.max(indicator)
        my_sheet.cell(r, 14 + 6*i).value = np.min(indicator)
        my_sheet.cell(r, 15 + 6*i).value = np.median(indicator)
        my_sheet.cell(r, 16 + 6*i).value = "No." + str(np.argmax(indicator) + 1)
        my_sheet.cell(r, 17 + 6*i).value = "No." + str(np.argmin(indicator) + 1)
    
    my_sheet.cell(r, 4).alignment = right_alignment

    my_wb.save(sheet_name)
    print("Save Successed!")

def evaluation(targetDir, *indicators) -> None:
    paretoFronts = sorted(glob.glob(targetDir + '/*.csv')) #list型に格納
    numOfSols = np.zeros(len(paretoFronts))
    cr = np.zeros(len(paretoFronts))
    
    if len(indicators) == 0:
        for t, pf in enumerate(paretoFronts):
            df = pd.read_csv(pf, index_col=0)
            solutions = df.values
            numOfSols[t] = solutions.shape[0]
            cr[t] = cover_rate(solutions, divNum=solutions.shape[0]) # 被覆率の計算
    else:
        numOfSols = indicators[0]
        cr = indicators[1]
    
    print("\n---Numerical Results---")
    print("THE NUMBER OF SOLS")
    display(paretoFronts, numOfSols) # display representative values
    
    print("\nCOVER RATE")
    display(paretoFronts, cr) # display representative values

def display(pfs, idx) -> None:
    print("Average is {}".format(np.average(idx)))
    print("Median  is {}".format(np.median(idx)))
    print("Maximum is {}, No.{} ({})".format(np.max(idx), np.argmax(idx)+1, pfs[np.argmax(idx)]))
    print("Minimum is {}, No.{} ({})".format(np.min(idx), np.argmin(idx)+1, pfs[np.argmin(idx)]))

def cover_rate(arcEval, divNum) -> float:
    K = arcEval.shape[1]
    min_f = np.array([np.min(arcEval[:, 0]), np.min(arcEval[:, 1])])
    max_f = np.array([np.max(arcEval[:, 0]), np.max(arcEval[:, 1])])

    width = np.array([max_f[0] - min_f[0], max_f[1] - min_f[1]]) / divNum

    region = np.zeros((K, divNum))
    cover = np.zeros(K)
    for k in range(K):
        for r in range(arcEval.shape[0]):
            for div in range(divNum):
                lowLim = min_f[k] + width[k] * div
                highLim = lowLim + width[k]

                if lowLim <= arcEval[r, k] and arcEval[r, k] <= highLim:
                    region[k, div] = region[k, div] + 1
                    break
        
        for div in range(divNum):
            if region[k, div] != 0:
                cover[k] = cover[k] + 1
        cover[k] = cover[k] / divNum
    return np.sum(cover) / K

def RNI(dFName1, dFName2) -> tuple:
    df1 = pd.read_csv(dFName1, index_col=0)
    df2 = pd.read_csv(dFName2, index_col=0)
    n1 = df1.values
    n2 = df2.values
    S1 = np.hstack((n1, np.zeros((n1.shape[0], 1))))
    S2 = np.hstack((n2, np.ones((n2.shape[0], 1))))
    S_union = np.vstack((S1, S2))
    #print(S_union.shape)
    S_front_eval = np.zeros((1000, 3))
    indices_sorted = np.argsort(S_union[:,0]) # f1軸に対して昇順ソート
    #print(indices_sorted)
    S_union_sorted = S_union[indices_sorted]
    #print(S_union_sorted)

    S_front_eval[0] = S_union_sorted[0]
    Na = 1
    n1_p = 0
    n2_p = 0
    for r in range(1, S_union_sorted.shape[0]):
        if S_union_sorted[r, 1] < S_front_eval[Na - 1, 1]:
            S_front_eval[Na] = S_union_sorted[r]
            if S_front_eval[Na, 2] == 0:
                n1_p += 1
            else:
                n2_p += 1
            Na += 1
    
    #print(S_front_eval.shape)
    #print(n1_p)
    #print(n2_p)
    return (n1_p / (n1_p + n2_p), n2_p / (n1_p + n2_p))

if __name__ == "__main__":
    args = sys.argv # コマンドライン引数

    try:
        option = args[1]    
    except IndexError:
        print("オプションを指定してください.\n(-rni: RNI計算, -val: 個数と被覆率計算)")
    else:
        if option == '-rni':
            pf1_name = input("1つ目のパレートフロントのファイル名を入力:")
            pf2_name = input("2つ目のパレートフロントのファイル名を入力:")
            print("【{}】と\n【{}】のRNIは...\n".format(pf1_name, pf2_name))
            rni = RNI(pf1_name, pf2_name)

            print("RNI = ", rni)
            if rni[0] > rni[1]:
                print("前者の方がイイ!")
            elif rni[0] < rni[1]:
                print("後者の方がイイ!")
            else:
                print("同率!")
        
        elif option == '-val':
            dir_name = input("ディレクトリ名を入力：")
            print("【{}】内のパレートフロントを評価".format(dir_name))
            try:
                evaluation(dir_name)
            except ValueError:
                print("ディレクトリ名に誤りがあります.")

        elif option == '-rniall':
            pf1_name = input("1つ目のパレートフロントのファイル名を入力:")
            numOfPF = 9

            pf2_list = []
            for i in range(numOfPF):
                pf = input("2-{}つ目のパレートフロントを入力: ".format(i+1))
                pf2_list.append(pf)

            for pf2_name in pf2_list:
                rni = RNI(pf1_name, pf2_name)
                print(rni)

        else:
            print("不正なオプションです.")